import re
from functools import wraps
from flask import render_template, request, jsonify, session, redirect, url_for, Response

from database import get_user_permissions
from . import bp
from . import model
from . import revenue

from datetime import datetime, date

from database import get_db, get_cursor


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


def get_perms():
    if session.get('is_admin'):
        return {
            'can_read': 1,
            'can_add': 1,
            'can_edit': 1,
            'can_delete': 1
        }
    return get_user_permissions(session.get('user_id'), 'RP02')


def _can_upload():
    perms = get_perms()
    return bool(perms.get('can_add') or perms.get('can_edit'))


def read_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        if not get_perms().get('can_read'):
            return jsonify({'error': 'No permission to read RP02'}), 403
        return f(*args, **kwargs)
    return decorated


def upload_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        if not _can_upload():
            return jsonify({'error': 'No permission to modify bill master data'}), 403
        return f(*args, **kwargs)
    return decorated


@bp.route('/module/RP02/')
@login_required
def index():
    perms = get_perms()
    if not perms.get('can_read'):
        return render_template('no_access.html'), 403
    return render_template('rp02.html', username=session.get('username'),
                           permissions=perms, can_upload=_can_upload())


@bp.route('/module/RP02/bill-master/')
@login_required
def bill_master_index():
    perms = get_perms()
    if not perms.get('can_read'):
        return render_template('no_access.html'), 403
    return render_template('bill_master.html', username=session.get('username'),
                           status=model.get_status(), permissions=perms,
                           can_upload=_can_upload())


@bp.route('/module/RP02/bill-master-report/')
@login_required
def bill_master_report_index():
    perms = get_perms()
    if not perms.get('can_read'):
        return render_template('no_access.html'), 403
    return render_template('bill_master_report.html',
                           username=session.get('username'),
                           permissions=perms)


@bp.route('/api/module/RP02/bill-master-report/data')
@read_required
def bill_master_report_data():
    return jsonify({'data': model.get_bill_master_report()})


@bp.route('/api/module/RP02/bill-master/template')
@read_required
def bill_master_template():
    return Response(
        model.build_template_csv(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename="RP02_bill_master_template.csv"'},
    )


@bp.route('/api/module/RP02/bill-master/preview', methods=['POST'])
@upload_required
def bill_master_preview():
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'No file provided'}), 400
    rows, errors = model.parse_upload(f)
    customers = model.get_customer_master()
    recon = model.reconcile(rows, customers)
    years = sorted({r['financial_year'] for r in rows if r.get('financial_year')})
    opts = {col: customers for col, info in recon.items() if info['unknown']}
    return jsonify({'total_rows': len(rows), 'format_errors': errors,
                    'years': years,
                    'reconciliation': recon,
                    'master_options': opts,
                    'addable_columns': ['customer_name']})


@bp.route('/api/module/RP02/bill-master/apply', methods=['POST'])
@upload_required
def bill_master_apply():
    import json as _json
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'No file provided'}), 400
    try:
        resolutions = _json.loads(request.form.get('resolutions') or '{}')
    except (ValueError, TypeError):
        resolutions = {}

    rows, errors = model.parse_upload(f)
    if errors:
        return jsonify({'error': 'Fix format errors before applying',
                        'format_errors': errors}), 400

    # 1) Add-to-master actions (customer master only).
    added, add_errors = [], []
    for value, res in (resolutions.get('customer_name') or {}).items():
        if isinstance(res, dict) and res.get('action') == 'add':
            try:
                if model.add_customer(value):
                    added.append({'column': 'customer_name', 'value': value})
            except Exception as e:  # noqa: BLE001 — surface, don't abort
                add_errors.append({'column': 'customer_name', 'value': value, 'error': str(e)})

    # 2) Replace actions rewrite the parsed rows, then per-FY replace insert.
    rows = model.apply_resolutions(rows, resolutions)
    inserted, years = model.replace_years(rows, session.get('user_id'))
    return jsonify({'inserted': inserted, 'years': years,
                    'added_to_master': added, 'add_errors': add_errors})


@bp.route('/api/module/RP02/bill-master/rows')
@read_required
def bill_master_rows():
    import json as _json
    try:
        page = int(request.args.get('page', 1))
        size = int(request.args.get('size', 50))
    except (TypeError, ValueError):
        page, size = 1, 50
    try:
        filters = _json.loads(request.args.get('colfilters') or '[]')
        if not isinstance(filters, list):
            filters = []
    except (ValueError, TypeError):
        filters = []
    rows, total = model.get_rows(page, size, filters)
    return jsonify({'data': rows, 'last_page': max(1, (total + size - 1) // size), 'total': total})


@bp.route('/api/module/RP02/bill-master/row/update', methods=['POST'])
@upload_required
def bill_master_row_update():
    data = request.json or {}
    if not data.get('id'):
        return jsonify({'error': 'Missing id'}), 400
    res = model.update_row(data['id'], data)
    if res.get('error'):
        return jsonify(res), 400
    return jsonify(res)


@bp.route('/api/module/RP02/bill-master/row/delete', methods=['POST'])
@login_required
def bill_master_row_delete():
    if not get_perms().get('can_delete'):
        return jsonify({'error': 'No permission to delete'}), 403
    data = request.json or {}
    if not data.get('id'):
        return jsonify({'error': 'Missing id'}), 400
    model.delete_row(data['id'])
    return jsonify({'success': True})


# ── Billing Pipeline dashboard ───────────────────────────────────────────────

@bp.route('/module/RP02/billing-dashboard/')
@read_required
def billing_dashboard_index():
    return render_template('billing_dashboard.html', username=session.get('username'))


@bp.route('/api/module/RP02/billing-dashboard/data')
@read_required
def billing_dashboard_data():
    return jsonify(model.get_billing_dashboard())


# Columns for the Excel dump — quantity and dates only, no rates or amounts.
_EXPORT_COLUMNS = [
    ('Status', 'status'),
    ('Vessel / MBC', 'vessel_name'),
    ('Material PO', 'material_po'),
    ('Customer', 'customer_name'),
    ('Type of Cargo', 'cargo_type'),
    ('Cargo', 'cargo_name'),
    ('Pending Qty (MT)', 'bl_qty'),
    ('Load Port', 'load_port'),
    ('MV/MBC', 'mv_mbc'),
    ('Discharge Commence', 'discharge_commence'),
    ('Discharge Completed', 'discharge_completed'),
    ('Age (days)', 'age_days'),
    ('Age Bucket', 'age_bucket'),
    ('Doc Status', 'doc_status'),
]


@bp.route('/api/module/RP02/billing-dashboard/export')
@read_required
def billing_dashboard_export():
    """Excel dump of the pending lines, plus a summary sheet of the same numbers
    the dashboard shows. Deliberately carries no rates or amounts.

    ?dim=&val=&scope= exports one breakdown bar's drilldown instead of the whole
    pipeline, filtered by the same model helper the modal uses."""
    import io
    from datetime import datetime as _dt
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    dim = request.args.get('dim')
    val = request.args.get('val')
    scope = request.args.get('scope', 'all')

    data = model.get_billing_dashboard()
    if dim or scope != 'all':
        rows, drill_desc = model.filter_pending(dim, val, scope, rows=data['rows'])
        data = {**data, 'rows': rows}
    else:
        drill_desc = None

    thin = Side(style='thin', color='000000')
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill('solid', fgColor='1E3A5F')
    hdr_font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
    cell_font = Font(name='Calibri', size=10)
    blocked_font = Font(name='Calibri', size=10, color='8A5A00')
    lft = Alignment(horizontal='left', vertical='center')
    rgt = Alignment(horizontal='right', vertical='center')
    ctr = Alignment(horizontal='center', vertical='center', wrap_text=True)

    wb = openpyxl.Workbook()

    # Sheet 1: the lines
    ws = wb.active
    ws.title = 'Pending Lines'
    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 22
    for i, (label, _k) in enumerate(_EXPORT_COLUMNS, 1):
        c = ws.cell(1, i, label)
        c.font = hdr_font; c.fill = hdr_fill; c.border = bdr; c.alignment = ctr
        ws.column_dimensions[get_column_letter(i)].width = max(12, len(label) + 3)

    numeric = {'bl_qty', 'age_days'}
    for row_i, r in enumerate(data['rows'], 2):
        blocked = r.get('status') != 'Ready to bill'
        # not `val` — that name holds the drilldown query param used below
        for col_i, (_label, key) in enumerate(_EXPORT_COLUMNS, 1):
            cell_val = r.get(key)
            c = ws.cell(row_i, col_i, '' if cell_val is None else cell_val)
            c.border = bdr
            c.font = blocked_font if blocked else cell_font
            c.alignment = rgt if key in numeric else lft
            if key == 'bl_qty' and cell_val is not None:
                c.number_format = '#,##0.00'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(_EXPORT_COLUMNS))}{max(len(data["rows"]) + 1, 1)}'

    # Sheet 2: the dashboard's own numbers, so the file stands alone
    ws2 = wb.create_sheet('Summary')
    ws2.column_dimensions['A'].width = 34
    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 14
    k = data['kpi']
    blocks = [
        ('Billing Pipeline — generated ' + data['generated_at'], None, None),
        ('Quantities are cargo tonnage (MT) and line counts only. No rates or amounts.', None, None),
    ]
    if drill_desc:
        # a drilldown export says what it is a slice of, so the file is not
        # mistaken for the whole pipeline once it leaves this screen
        blocks += [
            (None, None, None),
            ('Drilldown: ' + drill_desc, None, None),
            ('Lines in this export', len(data['rows']),
             round(sum(r['bl_qty'] or 0 for r in data['rows']), 2)),
        ]
    blocks += [
        (None, None, None),
        ('Whole pipeline', 'Lines', 'Qty (MT)'),
        ('Ready to bill', k['ready_lines'], k['ready_qty']),
        ('Blocked upstream', k['blocked_lines'], k['blocked_qty']),
        ('Oldest waiting (days)', k['oldest_days'], None),
        ('Customers waiting', k['customers'], None),
        ('Billed to date (bill master lines)', k['billed_lines'], None),
        (f"Billed in {k['month_label']}", k['billed_month'], None),
    ]
    for title, group in (('Ready to bill by age', data['ageing']),
                         ('Ready to bill by customer', data['by_customer']),
                         ('Ready to bill by cargo type', data['by_cargo']),
                         ('Blocked by document status', data['blocked_by_status'])):
        blocks += [(None, None, None), (title, 'Lines', 'Qty (MT)')]
        blocks += [(g['label'], g['lines'], g['qty']) for g in group]

    for row_i, (a, b, c_) in enumerate(blocks, 1):
        is_head = b in ('Lines',) or (row_i == 1)
        ca = ws2.cell(row_i, 1, a if a is not None else '')
        ca.font = Font(name='Calibri', size=10, bold=bool(is_head))
        ca.alignment = lft
        # not `val` — that name holds the drilldown query param used below
        for col_i, cell_val in ((2, b), (3, c_)):
            cc = ws2.cell(row_i, col_i, '' if cell_val is None else cell_val)
            cc.font = Font(name='Calibri', size=10, bold=bool(is_head))
            cc.alignment = rgt
            if col_i == 3 and isinstance(cell_val, (int, float)):
                cc.number_format = '#,##0.00'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    slug = ''
    if drill_desc:
        safe = re.sub(r'[^A-Za-z0-9]+', '_', (val or scope)).strip('_')[:40]
        slug = f'_{safe}' if safe else '_drilldown'
    fname = f"RP02_billing_pipeline{slug}_{_dt.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{fname}"'},
    )


#Revenue Regisrter


@bp.route('/module/RP02/revenue/')
@read_required
def revenue_report():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    return render_template(
        'rp02_revenue_report.html',
        username=session.get('username'),
        module_code='RP02'
    )


@bp.route('/api/module/RP02/revenue/data')
@read_required
def revenue_report_data():
    if 'user_id' not in session:
        return jsonify({'error': 'Not logged in'}), 401

    # Date range filters.
    # The UI can send from_date and to_date in YYYY-MM-DD format.
    # Both dates are inclusive. Existing month/year filtering is retained
    # for backward compatibility when date range is not supplied.
    from_date = request.args.get('from_date') or request.args.get('date_from')
    to_date = request.args.get('to_date') or request.args.get('date_to')
    month = request.args.get('month')
    year = request.args.get('year')

    conn = get_db()
    cur = get_cursor(conn)

    where = []
    params = []

    def _valid_date(value):
        if not value:
            return None
        try:
            return datetime.strptime(value, "%Y-%m-%d").date()
        except (TypeError, ValueError):
            return None

    from_date_obj = _valid_date(from_date)
    to_date_obj = _valid_date(to_date)

    # If both dates are selected, show only records between them,
    # including both the From Date and To Date.
    if from_date_obj and to_date_obj:
        if from_date_obj > to_date_obj:
            conn.close()
            return jsonify({
                'error': 'From Date cannot be later than To Date'
            }), 400

        where.append("ih.invoice_date::date BETWEEN %s AND %s")
        params.extend([from_date_obj, to_date_obj])

    elif from_date_obj:
        where.append("ih.invoice_date::date >= %s")
        params.append(from_date_obj)

    elif to_date_obj:
        where.append("ih.invoice_date::date <= %s")
        params.append(to_date_obj)

    # Keep the old month/year filters working if no date range is supplied.
    elif month:
        where.append("EXTRACT(MONTH FROM ih.invoice_date::date) = %s")
        params.append(int(month))

    if not from_date_obj and not to_date_obj and year:
        where.append("EXTRACT(YEAR FROM ih.invoice_date::date) = %s")
        params.append(int(year))

    where_sql = "WHERE " + " AND ".join(where) if where else ""

    cur.execute(f"""
SELECT
    ih.id,
    ih.invoice_number,
    ih.invoice_date,
    ih.customer_name,
    vc.sap_customer_code,
    ih.customer_gl_code,
    ih.customer_gstin,
    ih.sap_document_number,
    ih.subtotal AS basic_value,
    ih.sgst_amount,
    ih.cgst_amount,
    ih.igst_amount,
    ih.total_amount AS invoice_value,

    ih.gst_irn,
    ih.gst_ack_number,
    ih.gst_ack_date,

    il.rate,
    il.quantity,
    il.uom,
    il.sac_code,
    il.hsn_sac,
    il.gl_code,
    il.service_name,

    (
        SELECT fl.service_description
        FROM fdcn_lines fl
        WHERE TRIM(fl.gl_code) = TRIM(il.gl_code)
        ORDER BY fl.id DESC
        LIMIT 1
    ) AS grouping,

    il.cgst_rate,
    il.sgst_rate,
    il.igst_rate

FROM invoice_header ih

LEFT JOIN vessel_customers vc
    ON TRIM(LOWER(vc.name)) = TRIM(LOWER(ih.customer_name))

LEFT JOIN LATERAL (
    SELECT
        rate,
        quantity,
        uom,
        sac_code,
        hsn_sac,
        gl_code,
        service_name,
        cgst_rate,
        sgst_rate,
        igst_rate
    FROM invoice_lines
    WHERE invoice_id = ih.id
    ORDER BY
        (hsn_sac IS NOT NULL AND hsn_sac <> '') DESC,
        (sac_code IS NOT NULL AND sac_code <> '') DESC,
        id
    LIMIT 1
) il ON TRUE

{where_sql}

ORDER BY ih.invoice_date::date DESC NULLS LAST, ih.id DESC
""", params)

    rows = [dict(r) for r in cur.fetchall()]
    conn.close()

    out = []
    today = datetime.today().date()

    for r in rows:
        inv_date = r.get('invoice_date')

        basic = float(r.get('basic_value') or 0)
        sgst = float(r.get('sgst_amount') or 0)
        cgst = float(r.get('cgst_amount') or 0)
        igst = float(r.get('igst_amount') or 0)
        inv_val = float(r.get('invoice_value') or 0)

        cgst_rate = float(r.get('cgst_rate') or 0)
        sgst_rate = float(r.get('sgst_rate') or 0)
        igst_rate = float(r.get('igst_rate') or 0)

        tax_rate = igst_rate if igst_rate > 0 else (cgst_rate + sgst_rate)

        gl_code = str(r.get('gl_code') or '').strip()

        hsn_code = (
            r.get('sac_code')
            if gl_code == '4101071000'
            else (r.get('hsn_sac') or '')
        )

        tds_rate = revenue.TDS_PERCENT.get(gl_code, 0)
        tds_amount = round((basic * tds_rate) / 100, 2)
        net_receivable = round(inv_val - tds_amount, 2)

        if inv_date:
            if hasattr(inv_date, "year"):
                invoice_date = inv_date
            else:
                invoice_date = datetime.strptime(
                    str(inv_date)[:10], "%Y-%m-%d"
                ).date()

            days = (today - invoice_date).days
        else:
            days = ""
        
        bucket = revenue.age_bucket(days)

        out.append({
            'invoice_no': r.get('invoice_number') or '',
            'group_type': (
                'Group'
                if any(x in str(r.get('customer_name') or '').lower()
                       for x in ('jsw', 'amba'))
                else 'Non Group'
            ),
            'revenue_type_1': (
                'Scrap'
                if 'scrap' in str(r.get('grouping') or '').lower()
                else 'Operation'
            ),
            'revenue_type_2': (
                'Scrap'
                if 'scrap' in str(r.get('grouping') or '').lower()
                else 'Wharfage'
                if 'wharfage' in str(r.get('grouping') or '').lower()
                else 'Berth Hiri'
                if 'berth hire' in str(r.get('grouping') or '').lower()
                else 'Cargo Handling'
            ),
            'cargo_volume': '',

            'date': str(inv_date)[:10] if inv_date else '',

            'cust_code': r.get('sap_customer_code') or r.get('customer_gl_code') or '',
            'customer_name': r.get('customer_name') or '',
            'gl_code': gl_code,
            'grouping': r.get('grouping') or '',

            'qty': r.get('quantity'),
            'rate': r.get('rate'),

            'tax_category': (
                'IGST' if igst > 0
                else 'CGST+SGST' if (cgst > 0 or sgst > 0)
                else ''
            ),
            'tax_rate': tax_rate,

            'basic_value': basic,
            'sgst': sgst,
            'cgst': cgst,
            'igst': igst,
            'invoice_value': inv_val,

            'gstin': r.get('customer_gstin') or '',
            'sap_doc_no': r.get('sap_document_number') or '',

            'sac_code': '' if hsn_code else (r.get('sac_code') or ''),
            'hsn_code': hsn_code,
            'irn': r.get('gst_irn') or '',
            'irn_date': (
                r.get('gst_ack_date').strftime('%Y-%m-%d')
                if r.get('gst_ack_date')
                else ''
            ),
            'ack_date': (
                r.get('gst_ack_date').strftime('%Y-%m-%d')
                if r.get('gst_ack_date')
                else ''
            ),
            'ack_no': r.get('gst_ack_number') or '',

            'tds_tcs': tds_amount,
            'net_receivable': net_receivable,
            'days': days,
            'bucket': bucket,
            'source': 'Live',
        })

    # Rows invoiced before go-live live in the backdated upload, not in
    # invoice_header — the register is the two put together.
    # Backdated rows must use the same selected date range as live rows.
    # The existing helper is month/year based, so fetch the relevant rows
    # first and then apply the inclusive date-range filter to the combined
    # register. This keeps both Live and Backdated data consistent.
    out.extend(revenue.get_register_rows(month, year))

    if from_date_obj or to_date_obj:
        def _row_date(row):
            value = row.get('date')
            if not value:
                return None
            try:
                return datetime.strptime(str(value)[:10], "%Y-%m-%d").date()
            except (TypeError, ValueError):
                return None

        filtered_out = []
        for row in out:
            row_date = _row_date(row)
            if row_date is None:
                continue
            if from_date_obj and row_date < from_date_obj:
                continue
            if to_date_obj and row_date > to_date_obj:
                continue
            filtered_out.append(row)
        out = filtered_out

    out.sort(key=lambda r: (r.get('date') or ''), reverse=True)
    return jsonify({
        'data': out,
        'from_date': from_date,
        'to_date': to_date
    })

# ── Revenue Register — backdated upload ─────────────────────────────────────

@bp.route('/module/RP02/revenue-backdated/')
@read_required
def revenue_backdated_index():
    return render_template('revenue_backdated.html', username=session.get('username'),
                           status=revenue.get_status(), can_upload=_can_upload(),
                           fields=revenue.FIELDS, derived=revenue.DERIVED_FIELDS)

@bp.route('/module/RP02/cargo-report/')
@login_required
def cargo_report_rp02():
    return render_template(
        'cargo_report/cargo_report.html',
        username=session.get('username'),
        module_code='RP02',
        module_href='/module/RP02/'
    )


@bp.route('/api/module/RP02/revenue-backdated/template')
@read_required
def revenue_backdated_template():
    return Response(
        revenue.build_template_csv(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename="RP02_revenue_register_template.csv"',
                 # the columns change as the register sheet does — never serve a cached copy
                 'Cache-Control': 'no-store'},
    )


@bp.route('/api/module/RP02/revenue-backdated/preview', methods=['POST'])
@upload_required
def revenue_backdated_preview():
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'No file provided'}), 400
    rows, errors = revenue.parse_upload(f)
    months = sorted({revenue.month_key(r['invoice_date']) for r in rows})
    return jsonify({'total_rows': len(rows), 'format_errors': errors, 'months': months,
                    'sample': [dict({k: str(v) if v is not None else ''
                                     for k, v in r.items()},
                                    **dict(zip(('days', 'bucket'),
                                               map(str, revenue.derive_ageing(r['invoice_date'])))))
                               for r in rows[:5]]})


@bp.route('/api/module/RP02/revenue-backdated/apply', methods=['POST'])
@upload_required
def revenue_backdated_apply():
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'No file provided'}), 400
    rows, errors = revenue.parse_upload(f)
    if errors:
        return jsonify({'error': 'Fix format errors before applying',
                        'format_errors': errors}), 400
    inserted, months = revenue.replace_months(rows, session.get('user_id'))
    return jsonify({'inserted': inserted, 'months': months})


@bp.route('/api/module/RP02/revenue-backdated/rows')
@read_required
def revenue_backdated_rows():
    import json as _json
    try:
        page = int(request.args.get('page', 1))
        size = int(request.args.get('size', 50))
    except (TypeError, ValueError):
        page, size = 1, 50
    try:
        filters = _json.loads(request.args.get('colfilters') or '[]')
        if not isinstance(filters, list):
            filters = []
    except (ValueError, TypeError):
        filters = []
    rows, total = revenue.get_rows(page, size, filters)
    return jsonify({'data': rows, 'last_page': max(1, (total + size - 1) // size),
                    'total': total})


@bp.route('/api/module/RP02/revenue-backdated/row/update', methods=['POST'])
@upload_required
def revenue_backdated_row_update():
    data = request.json or {}
    if not data.get('id'):
        return jsonify({'error': 'Missing id'}), 400
    res = revenue.update_row(data['id'], data)
    if res.get('error'):
        return jsonify(res), 400
    return jsonify(res)


@bp.route('/api/module/RP02/revenue-backdated/row/delete', methods=['POST'])
@login_required
def revenue_backdated_row_delete():
    if not get_perms().get('can_delete'):
        return jsonify({'error': 'No permission to delete'}), 403
    data = request.json or {}
    if not data.get('id'):
        return jsonify({'error': 'Missing id'}), 400
    revenue.delete_row(data['id'])
    return jsonify({'success': True})
